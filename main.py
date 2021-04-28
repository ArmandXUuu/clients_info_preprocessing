import csv
import re
import sys
from openpyxl import load_workbook, Workbook
import finish

name_style = 'pure'

class Client:
    def __init__(self):
        self.name = None
        self.number = 0
        self.communications = []
        self.gender = None
        self.suspecious_communications = []
        self.identity = ""
        self.buffer1 = None
        self.buffer2 = None

    def communications_to_string(self):
        tmp = []
        for com in self.communications:
            tmp.append(" ".join(com))
        return '; '.join(tmp)

    def susp_communications_to_string(self):
        tmp = []
        for com in self.suspecious_communications:
            tmp.append(" ".join(com))
        return ''.join(tmp)

def process_gender(row):
    MR_pattern = re.compile(r'^.* MR .*$')
    MS_pattern = re.compile(r'^.* MS .*$')
    MRS_pattern = re.compile(r'^.* MRS .*$')
    if MR_pattern.match(row) != None:
        return "MR"
    if MS_pattern.match(row) != None:
        return "MS"
    if MRS_pattern.match(row) != None:
        return "MRS"
    return None

def process_identity(row):
    SD_pattern = re.compile(r'^.* SD .*$')
    CHD_pattern = re.compile(r'^.* CHD .*$')
    INF_pattern = re.compile(r'^.* INF .*$')
    if SD_pattern.match(row) != None:
        return "SD"
    if CHD_pattern.match(row) != None:
        return "CHD"
    if INF_pattern.match(row) != None:
        return "INF"
    return None

def collect_files():
    if len(sys.argv) - 2 == 0 or sys.argv[1] == '-n':
        file_names = ['input.csv']
    else:
        file_names = sys.argv[2::]
    return len(sys.argv) - 2, file_names

def read_from_xl(input_file = 'test.xlsx'):
    wb = load_workbook(input_file)
    first_column = wb[wb.sheetnames[0]]['A']
    rows = []
    for r in range(len(first_column)):
        if type(first_column[r].value) == int:
            pre_string = str(first_column[r].value)
        else:
            pre_string = first_column[r].value
        row = re.sub(' +', ' ', pre_string.lstrip())

        rows.append(row)
    return rows

def write_to_xl(clients, output_file = 'output.xlsx'):
    output = Workbook()
    sheet = output.active

    sheet['A1'] = 'number'
    sheet['B1'] = 'gender'
    sheet['C1'] = 'name'
    sheet['D1'] = 'buffer1'
    sheet['E1'] = 'buffer2'
    sheet['F1'] = 'identity'
    sheet['G1'] = 'contact'
    sheet['H1'] = 'suspecious_contact'
    i = 2

    for client in clients:
        sheet['A'+str(i)] = client.number
        sheet['B'+str(i)] = client.gender
        sheet['C'+str(i)] = client.name
        sheet['D'+str(i)] = client.buffer1
        sheet['E'+str(i)] = client.buffer2
        sheet['F'+str(i)] = client.identity
        sheet['G'+str(i)] = client.communications_to_string()
        sheet['H'+str(i)] = client.susp_communications_to_string()
        i += 1
    
    output.save(output_file)

def process(input_file = "input.xlsx", output_file = "output.xlsx"):
    rows = read_from_xl(input_file)

    start_person = re.compile(r'^\d\d\d .*')
    start_person_Car = re.compile(r'^\d\d\d[A-Z] .*')
    start_OSI = re.compile(r'^OSI.*')
    pure_numbers = re.compile(r'^\d.*')
    six_reservation_code = re.compile(r'[A-Za-z0-9]{6}')
    ends_with_SD = re.compile(r'.*SD$')
    
    row = None
    clients = []
    client_tmp = Client()

    for row in rows:
        row_tmp = str.split(row, " ")
        # when a new person comes
        if start_person.match(row) != None or start_person_Car.match(row) != None: 
            clients.append(client_tmp)
            client_tmp = Client()

            client_tmp.gender = process_gender(row)
            client_tmp.identity = process_identity(row)

            client_tmp.number = row_tmp[0]

            # client.name & client.buffer
            client_tmp.name = row_tmp[1]
            # an exception when SD is just behind his/her name
            if sys.argv[1] == '-ib' and ends_with_SD.match(client_tmp.name):
                print("\n" + client_tmp.name)
                ends_SD = input("Is this name ends with an SD, which is not correct ? (y/n) ")
                if ends_SD == 'y':
                    client_tmp.name = client_tmp.name.rstrip("SD")
                    client_tmp.identity = "SD"
            # try to correct clients' names (an espace between them)
            if sys.argv[1] == '-ib' and len(row_tmp[2]) != 6 and (row_tmp[2] not in ["MR", "MRS", "CHD", "INF", "SD", "MS"]):
                if row_tmp[2] == '+':
                    espace_exists = "y"
                else:
                    print("\n" + " ".join(row_tmp[0:2:]), "< " + row_tmp[2] + " >", " ".join(row_tmp[3::]))
                    espace_exists = input("Is there an espace in his/her name ? (y/n) ")
                if espace_exists == "y":
                    client_tmp.name = client_tmp.name + row_tmp[2]
                    client_tmp.buffer1 = row_tmp[3]
                    client_tmp.buffer2 = row_tmp[4]
            else:
                client_tmp.buffer1 = row_tmp[2]
                client_tmp.buffer2 = row_tmp[3]
            # change the name output style
            if name_style == 'pure':
                client_tmp.name = client_tmp.name.lstrip('1').lstrip('2').replace('/', '')

        # still adding information to this client_tmp
        else :
            if start_OSI.match(row) != None:
                client_tmp.communications.append(row_tmp[2::])
            elif (pure_numbers.match(row) != None):
                client_tmp.suspecious_communications = row

    clients.append(client_tmp)
    write_to_xl(clients, output_file)
    finish.finish(output_file)

def init():
    global name_style
    print("Please choose a sytle for name output:")
    print("\t1. 1ZHANG/SAN\t2. ZHANGSAN")
    inp = input("Please enter 1 or 2: ")
    if inp == "1":
        name_style = 'complex'

def main():
    init()
    file_number, file_names = collect_files()
    for file_name in file_names:
        file_name_tmp = file_name.rstrip('.xlsx')
        process(file_name, file_name_tmp + '_out.xlsx')
    
    print("\nThank you for using, processing complete. " + str(file_number) + " files procesed.")

if __name__ == '__main__':
    main()